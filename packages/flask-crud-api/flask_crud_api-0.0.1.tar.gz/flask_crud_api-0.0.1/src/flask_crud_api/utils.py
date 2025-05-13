import os
import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side


def str2datetime(strs, format="%Y-%m-%d %H:%M:%S"):
    return datetime.datetime.strptime(strs, format)


def datetime2str(_datetime, format="%Y-%m-%d %H:%M:%S"):
    return _datetime.strftime(format)


class Excel:
    def __init__(self, filename, headers=None, width=15):
        self.filename = filename
        self.headers = headers
        self.width = width
        self.wb: openpyxl.Workbook = None
        if self.headers is None:
            self.headers = []
        self.highlight = NamedStyle(name="highlight")
        self.highlight.font = Font(bold=False, size=13.5, color="000000")
        bd = Side(style="thin", color="000000")
        self.highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self._line = 1

    @classmethod
    def from_write_excel(cls, filename, headers, width=15):
        obj = cls(filename, headers, width)
        obj.wb = openpyxl.Workbook()
        return obj

    @classmethod
    def from_read_excel(cls, filename):
        obj = cls(filename)
        obj.wb = openpyxl.load_workbook(filename)
        return obj

    def write(self, lines):
        sheet = self.wb.active
        for idx in range(len(self.headers)):
            sheet.column_dimensions[get_column_letter(idx + 1)].width = self.width
        for _, line in enumerate(lines):
            for idx, column in enumerate(line):
                space = f"{get_column_letter(idx+1)}{self._line}"
                sheet[space] = column
                sheet[space].style = self.highlight
            self._line += 1

    def write_headers(self):
        self.write([self.headers])

    def save(self):
        name, ext = os.path.splitext(self.filename)
        ext = ext or ".xlsx"
        filepath = (
            f"./static/{name}-{datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')}{ext}"
        )
        self.wb.save(filepath)
        return filepath

    def read(self):
        sheet = self.wb.active
        for row in sheet.iter_rows(values_only=True):
            yield row
