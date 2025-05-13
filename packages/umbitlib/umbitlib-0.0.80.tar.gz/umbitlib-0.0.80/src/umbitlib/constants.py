import os
from openpyxl import load_workbook

# List of Columns with categorical variables
LST_CATG_VARIABLES = ['location', 'billing_provider', 'proc_code', 'pos_type', 'pos_group']


####################################################################################
##  Postgres Database Credentials
####################################################################################

# First, Try to pull credentials from the system environment variables
try:
    # Development Server
    PG_DEV_DB = os.environ['PG_DEV_DB']
    PG_DEV_HOST = os.environ['PG_DEV_HOST']
    PG_DEV_PASS = os.environ['PG_DEV_PASS']
    PG_DEV_PORT = os.environ['PG_DEV_PORT']
    PG_DEV_USER = os.environ['PG_DEV_USER']

    # Production Server
    PG_PROD_DB = os.environ['PG_PROD_DB']
    PG_PROD_HOST = os.environ['PG_PROD_HOST']
    PG_PROD_PASS = os.environ['PG_PROD_PASS']
    PG_PROD_PORT = os.environ['PG_PROD_PORT']
    PG_PROD_USER = os.environ['PG_PROD_USER']
    
# If that doesn't work try the box drive
except:
    try:
        # Box Drive Excel File
        data_source_file_postgres = r"C:\Cloud\Box\UMB_DataScience\System Environment Variables.xlsx"
        wb_credentials_postgres = load_workbook(data_source_file_postgres)
        sheet_postgres = wb_credentials_postgres['Sheet1']
        
        # Development Server
        PG_DEV_DB = sheet_postgres['B5'].value
        PG_DEV_HOST = sheet_postgres['B6'].value
        PG_DEV_PASS = sheet_postgres['B7'].value
        PG_DEV_PORT = sheet_postgres['B8'].value
        PG_DEV_USER = sheet_postgres['B9'].value

        # Production Server
        PG_PROD_DB = sheet_postgres['B11'].value
        PG_PROD_HOST = sheet_postgres['B12'].value
        PG_PROD_PASS = sheet_postgres['B13'].value
        PG_PROD_PORT = sheet_postgres['B14'].value
        PG_PROD_USER = sheet_postgres['B15'].value
        
    # If that doesn't work try to pull them from the G: Drive Excel File    
    except:        
        try:
            # G: Drive Excel File
            username = os.getlogin()
            data_source_file_postgres = f"//ad.utah.edu/uuhc/users/{username}/Data Sources/Postgres.xlsx"
            wb_credentials_postgres = load_workbook(data_source_file_postgres)
            sheet_postgres = wb_credentials_postgres['Sheet1']

            # Development Server
            PG_DEV_DB = sheet_postgres['B1'].value
            PG_DEV_HOST = sheet_postgres['B2'].value
            PG_DEV_PASS = sheet_postgres['B3'].value
            PG_DEV_PORT = sheet_postgres['B4'].value
            PG_DEV_USER = sheet_postgres['B5'].value

            # Production Server
            PG_PROD_DB = sheet_postgres['B7'].value
            PG_PROD_HOST = sheet_postgres['B8'].value
            PG_PROD_PASS = sheet_postgres['B9'].value
            PG_PROD_PORT = sheet_postgres['B10'].value
            PG_PROD_USER = sheet_postgres['B11'].value
            
        # Finally if none of those work set them all to a blank string so the Library doesn't error when imported
        except:
            # Development Server
            PG_DEV_DB = ''
            PG_DEV_HOST = ''
            PG_DEV_PASS = ''
            PG_DEV_PORT = ''
            PG_DEV_USER = ''

            # Production Server
            PG_PROD_DB = ''
            PG_PROD_HOST = ''
            PG_PROD_PASS = ''
            PG_PROD_PORT = ''
            PG_PROD_USER = ''


####################################################################################
##  Oracle Database Credentials
####################################################################################
username = os.getlogin()
data_source_file_oracle = f"//ad.utah.edu/uuhc/users/{username}/Data Sources/UIDPWD.xlsx"
wb_credentials_oracle = load_workbook(data_source_file_oracle)
sheet_oracle = wb_credentials_oracle['Sheet1']
ORACLE_USER = sheet_oracle['A2'].value
ORACLE_PASS = sheet_oracle['B2'].value
# %%
