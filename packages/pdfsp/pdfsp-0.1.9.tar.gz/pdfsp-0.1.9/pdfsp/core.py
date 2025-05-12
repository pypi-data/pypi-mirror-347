# This file is part of the pdfsp project
# Copyright (C) 2025 Sermet Pekin
#
# This source code is free software; you can redistribute it and/or
# modify it under the terms of the European Union Public License
# (EUPL), Version 1.2, as published by the European Commission.
#
# You should have received a copy of the EUPL version 1.2 along with this
# program. If not, you can obtain it at:
# <https://joinup.ec.europa.eu/collection/eupl/eupl-text-eupl-12>.
#
# This source code is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# European Union Public License for more details.
#
# Alternatively, if agreed upon, you may use this code under any later
# version of the EUPL published by the European Commission.
import pdfplumber
from dataclasses import dataclass
import os
import pandas as pd

from ._typing import (
    T_Path,
    T_OptionalPath,
    Generator,
    Path,
    T_pandas_df,
    T_List_path,
    T_List_str,
    Dict,
)

from collections import Counter
from ._options import Options


@dataclass
class DataFrame:
    df: T_pandas_df
    path: T_Path
    out: T_OptionalPath = None
    page: int = 1
    index: int = 1
    extra: tuple = ()
    name: str = ""

    def __post_init__(self):
        if self.out is None:
            self.out = "Output"
        self.out = Path(self.out)
        self.df = self.make_unique_cols(self.df)
        self.name = Path(self.path).stem.split(".pdf")[0]

    def make_unique_cols(self, df: T_pandas_df) -> T_pandas_df:
        cols = [str(x) for x in df.columns]
        df.columns = self.make_unique(cols)
        return df

    def make_unique(self, cols: T_List_str) -> T_List_str:
        counter = Counter()
        unique_cols = []
        for col in cols:
            counter[col] += 1
            suffix = f"-{counter[col]-1}" if counter[col] > 1 else ""
            unique_cols.append(f"{col}{suffix}")
        return unique_cols

    def get_file_name(self) -> str:
        return f"[{self.name}]-Page {self.page}-T {self.index}.xlsx"

    def create_dir(self) -> None:
        os.makedirs(self.out, exist_ok=True)

    def write(self) -> None:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        self.create_dir()
        file_name = self.get_file_name()
        wb = Workbook()
        ws = wb.active
        title = f"{self.name}-Table-{self.index}"
        ws.title = title
        for r_idx, row in enumerate(
            dataframe_to_rows(self.df, index=False, header=True), start=1
        ):
            ws.append(row)
        footnote_row = len(self.df) + 4
        ws.cell(row=footnote_row, column=1, value=f"Footnote: {title} ")
        paragraph_row = footnote_row + 4
        ws.cell(
            row=paragraph_row,
            column=1,
            value=f"This table was extracted from {self.path} with pdfsp package.",
        )
        wb.save(self.out / file_name)
        print(f"[writing table] {file_name}")


def check_folder(folder: T_Path) -> bool:
    """Check if the folder exists and is a directory."""
    folder = Path(folder)
    if not folder.exists():
        print(f"Folder `{folder}` does not exist.")
        return False
    if not folder.is_dir():
        print(f"`{folder}` is not a directory.")
        return False
    return True


class Folder(object):
    def __init__(self, folder: T_OptionalPath = None):
        if folder is None:
            folder = "."
        self.folder = Path(folder)
        if not check_folder(self.folder):
            raise ValueError(f"Invalid folder: {folder}")

    def __str__(self) -> str:
        return str(self.folder)

    def __repr__(self) -> str:
        return f"Folder({self.folder})"

    def __iter__(self):
        """Iterate over all files in the folder."""
        for file in self.folder.iterdir():
            if file.is_file() and file.suffix.lower() == ".pdf":
                yield file
            else:
                ...
                # print(f"Skipping non-PDF file: {file}")

    def __len__(self) -> int:
        """Get the number of PDF files in the folder."""
        return len(list(self.folder.glob("*.pdf")) + list(self.folder.glob("*.PDF")))


def extract_tables_from_pdf(
    pdf_path: T_Path, out: T_OptionalPath = None
) -> Generator[DataFrame, None, None]:
    """Extract tables from a PDF file and yield DataFrame objects."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"Extracting tables from `{pdf_path}`")
            for i, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                for index, table in enumerate(tables):
                    try:
                        df = pd.DataFrame(table[1:], columns=table[0])
                        yield DataFrame(df, pdf_path, out, page=i, index=index + 1)
                    except Exception as e:
                        print(
                            f"⚠️ Failed to process table {index + 1} on page {i} of `{pdf_path}`: {e}"
                        )
    except Exception as e:
        print(f"❌ Failed to extract tables from `{pdf_path}`: {e}")

from typing import List  

def _strip_repeated_header(df: pd.DataFrame) -> pd.DataFrame:
    """Remove the first row if it repeats the header (common in continuation pages)."""
    if df.empty:
        return df

    # Convert all to string to avoid type mismatches during comparison
    header_as_list = list(map(str, df.columns))
    first_row = list(map(str, df.iloc[0].tolist()))

    if header_as_list == first_row:
        return df.iloc[1:].reset_index(drop=True)
    return df

from typing import List

def combine_tables_by_continuation(dfs: List[DataFrame]) -> List[pd.DataFrame]:
    """
    Merge DataFrames that belong to the same logical table
    (identified by identical column names) and strip any
    repeated header rows that appeared after a page break.
    """
    if not dfs:
        return []

    combined, current_parts = [], [_strip_repeated_header(dfs[0].df)]

    for prev, nxt in zip(dfs, dfs[1:]):
        same_header = list(prev.df.columns) == list(nxt.df.columns)
        if same_header:
            current_parts.append(_strip_repeated_header(nxt.df))
        else:
            combined.append(pd.concat(current_parts, ignore_index=True))
            current_parts = [_strip_repeated_header(nxt.df)]

    # last group
    combined.append(pd.concat(current_parts, ignore_index=True  ))
    return combined


def write_combined_tables(dfs: List[DataFrame], pdf_path: T_Path, out: T_OptionalPath = None) -> None:
    """Write combined tables to separate Excel files."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    out = Path(out or ".")
    out.mkdir(exist_ok=True)

    combined_tables = combine_tables_by_continuation(dfs)

    for idx, combined_df in enumerate(combined_tables, start=1):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Table {idx}"

        for row in dataframe_to_rows(combined_df, index=False, header=True):
            ws.append(row)

        file_name = f"{Path(pdf_path).stem}-Table-{idx}.xlsx"
        wb.save(out / file_name)
        print(f"[writing combined table] {file_name}")



def process_folder_combine(options: Options) -> Dict[str, Dict[str, int]]:
    """
    Process all PDF files in a folder and return a report of successes, 
    failures, and the number of combined tables extracted per file.
    """
    report = {"success": {}, "failed": []}  # success: filename → table_count
    _folder = Folder(options.source_folder)

    for file in _folder:
        dfs = []
        try:
            for _df in extract_tables_from_pdf(file, options.output_folder):
                if _df:
                    dfs.append(_df)
        except Exception as e:
            print(f"❌ Error processing `{file}`: {e}")
            report["failed"].append(str(file))
            continue

        if dfs:
            combined_tables = combine_tables_by_continuation(dfs)
            write_combined_tables(dfs, file, options.output_folder)
            report["success"][str(file)] = len(combined_tables)
        else:
            print(f"⚠️ No tables found in `{file}`")
            report["failed"].append(str(file))

    print_summary_report(report)
    return report



def process_folder(options: Options) -> Dict[str, Dict[str, int]]:
    """Process all PDF files in a folder and return a report of successes, failures, and extracted table counts."""
    if options.combine:
        return process_folder_combine(options)
    report = {"success": {}, "failed": []}  # filename: table_count
    _folder = Folder(options.source_folder)
    for file in _folder:
        table_count = 0
        try:
            for _df in extract_tables_from_pdf(file, options.output_folder):
                if _df:
                    _df.write()
                    table_count += 1
            if table_count > 0:
                report["success"][str(file)] = table_count
            else:
                print(f"⚠️ No tables found in `{file}`")
                report["failed"].append(str(file))
        except Exception as e:
            print(f"❌ Error processing `{file}`: {e}")
            report["failed"].append(str(file))

    print_summary_report(report)
    return report


def print_summary_report(report: Dict[str, Dict[str, int]]) -> None:
    """Print a summary report including how many tables were extracted per file."""
    print("\n=== 📊 Extraction Summary Report ===")

    success_files = report["success"]
    failed_files = report["failed"]

    print(f"✅ Successful Files: {len(success_files)}")
    for file, count in success_files.items():
        print(f"   - {file} → 🗂️ {count} tables extracted")

    print(f"\n❌ Failed Files: {len(failed_files)}")
    for f in failed_files:
        print(f"   - {f}")

    if not failed_files:
        print("\n🎉 All files processed successfully!")
    else:
        print("\n⚠️ Some files failed to process. See details above.")


def extract_tables(options: Options) -> None:
    """Extract tables from all PDF files in the specified folder."""
    process_folder(options)
    print("Extraction completed.")
