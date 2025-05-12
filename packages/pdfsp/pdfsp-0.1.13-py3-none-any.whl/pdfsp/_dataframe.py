from dataclasses import dataclass
import os

from ._typing import (
    T_Path,
    T_OptionalPath,
    Path,
    T_pandas_df,
    T_List_str,
)

from collections import Counter

# ................................................................

# ................................................................


@dataclass
class DataFrame:
    df: T_pandas_df
    path: T_Path
    out: T_OptionalPath = None
    page: int = 1
    index: int = 1
    extra: tuple = ()
    name: str = ""

    def __post_init__(self):
        if self.out is None:
            self.out = "Output"
        self.out = Path(self.out)
        self.df = self.make_unique_cols(self.df)
        self.name = Path(self.path).stem.split(".pdf")[0]

    def make_unique_cols(self, df: T_pandas_df) -> T_pandas_df:
        cols = [str(x) for x in df.columns]
        df.columns = self.make_unique(cols)
        return df

    def make_unique(self, cols: T_List_str) -> T_List_str:
        counter = Counter()
        unique_cols = []
        for col in cols:
            counter[col] += 1
            suffix = f"-{counter[col]-1}" if counter[col] > 1 else ""
            unique_cols.append(f"{col}{suffix}")
        return unique_cols

    def get_file_name(self) -> str:
        return f"[{self.name}]-Page {self.page}-T {self.index}.xlsx"

    def create_dir(self) -> None:
        os.makedirs(self.out, exist_ok=True)

    def write(self) -> None:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        self.create_dir()
        file_name = self.get_file_name()
        wb = Workbook()
        ws = wb.active
        title = f"{self.name}-Table-{self.index}"
        ws.title = title
        for r_idx, row in enumerate(
            dataframe_to_rows(self.df, index=False, header=True), start=1
        ):
            ws.append(row)
        footnote_row = len(self.df) + 4
        ws.cell(row=footnote_row, column=1, value=f"Footnote: {title} ")
        paragraph_row = footnote_row + 4
        ws.cell(
            row=paragraph_row,
            column=1,
            value=f"This table was extracted from {self.path} with pdfsp package.",
        )
        wb.save(self.out / file_name)
        print(f"[writing table] {file_name}")
