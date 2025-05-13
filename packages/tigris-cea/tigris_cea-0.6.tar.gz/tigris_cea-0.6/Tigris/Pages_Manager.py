# -*- coding: utf-8 -*-

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
import os
import numpy as np
import openpyxl
import time
#------------------------------------------------------------------------------
from PyQt5 import uic
from PyQt5.QtCore import Qt, QRect, pyqtSignal
from PyQt5.QtGui import QCursor
from PyQt5.QtWidgets import QDialog, QComboBox, QSizePolicy, QTabBar, QMenu, QTabWidget, QFileDialog, QLineEdit
#------------------------------------------------------------------------------
from Page import Page, Page_Config
from Graph import Graph_Config
from utils_qt import get_icon, question_box, clear_layout
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
class Pages_Manager :
    
    #==========================================================================
    def __init__(self, mw) :
        #----------------------------------------------------------------------
        self.mw = mw
        
        #----------------------------------------------------------------------
        self.page_configs  = Page_Config.get_configs()
        self.graph_configs = Graph_Config.get_configs()
        
        #----------------------------------------------------------------------
        self.new_page_dialog = Page_Dialog(self)
        self.tab_widget = Pages_TabWidget(self)
        
        #----------------------------------------------------------------------
        self.pages = []
        
    #==========================================================================
    def get_page(self, page_index=None, tab_index=None, name=None) :
        #----------------------------------------------------------------------
        if tab_index is not None : return self.tab_widget.widget(tab_index)
        for page in self.pages :
            if page_index is not None and page.index == page_index : return page
            if name is not None and page.name == name : return page
                
    #==========================================================================
    def get_index(self, page) : return self.tab_widget.indexOf(page)
    
    #==========================================================================
    def get_graph_page(self, graph) :
        #----------------------------------------------------------------------
        for page in self.pages :
            if page.graphs is None : continue
            if graph in page.graphs :
                return page
        #----------------------------------------------------------------------
        return None
        
    #==========================================================================
    def open_new_page_dialog(self) :
        #----------------------------------------------------------------------
        self.new_page_dialog.configure()
        self.new_page_dialog.show()
        
        #----------------------------------------------------------------------
        return self.new_page_dialog
        
    #==========================================================================
    def new_page(self, name, page_config, graph_configs) :
        #----------------------------------------------------------------------
        return Page(self, name, page_config, graph_configs)
    
    #==========================================================================
    def save_image(self) :
        #----------------------------------------------------------------------
        if len(self.pages) == 0 : return
        #----------------------------------------------------------------------
        self.tab_widget.currentWidget().save_image()
        
    #==========================================================================
    def save_data(self, filepath=None, page=None, graph=None) :
        #----------------------------------------------------------------------
        if len(self.pages) == 0 :
            return
        
        #----------------------------------------------------------------------
        if filepath is None :
            defaultname = "export_data.xlsx"
            if len(self.mw.files_manager.files) == 1 : defaultname = os.path.join(os.path.dirname(self.mw.files_manager.files[0].nc_path), defaultname)
            filepath,_ = QFileDialog.getSaveFileName(self.mw, "Enregistrer sous", defaultname, "Tous les fichiers (*);;Excel (*.xlsx)", "Excel (*.xlsx)")
            
        #----------------------------------------------------------------------
        if filepath == '' :
            return
        
        #----------------------------------------------------------------------
        print("> Sauvegarde des données ... ", end="", flush=True)
        
        #----------------------------------------------------------------------
        infos = []
        blocks = {} # [sheet_name] = [block1, block2, ...]
        for _page in self.pages :
            if page is not None and page.index != _page.index : continue
            for _graph in _page.graphs :
                if graph is not None and graph.index != _graph.index : continue
                
                _blocks = _graph._get_save_data()
                for sheet_name,_block in _blocks.items() :
                    if sheet_name not in blocks : blocks[sheet_name] = []
                    blocks[sheet_name] += _block
                    for b in _block :
                        for info in b['infos'] :
                            info['sheet'] = sheet_name
                            infos.append(info)
        
        #----------------------------------------------------------------------
        if len(blocks.keys()) == 0 :
            print("OK (aucune donnée)")
            return
        
        #----------------------------------------------------------------------
        wb = openpyxl.Workbook()
        for sheet_name in wb.sheetnames : del wb[sheet_name]
        
        #----------------------------------------------------------------------
        dims = []
        for info in infos : dims += [dim for dim in info['dimensions'] if dim not in dims]
        dims.sort()
        for dim in 'ZYX' :
            if dim in dims :
                dims.remove(dim)
                dims.insert(0, dim)
        
        #----------------------------------------------------------------------
        ws = wb.create_sheet(title='Informations')
        heads = ['Feuille','Balise','Fichier','Chemin','Dimensions','Indices']
        heads += ['Coord_{}'.format(dim) for dim in dims]
        
        for c,field in enumerate(heads) : ws.cell(row=1, column=c+1, value=field)
        for r,info in enumerate(infos) :
            element = info['element']
            line = []
            line.append(info['sheet'])
            line.append(info['beacon'])
            line.append(element.file.alias)
            line.append('/'.join(info['path']))
            line.append(",".join(info['dimensions']))
            line.append(",".join(info['indices']))
            line += [info['coordinates'].get(dim,'').format(dim) for dim in dims]
            for c,field in enumerate(line) : ws.cell(row=r+2, column=c+1, value=field)
            
        #----------------------------------------------------------------------
        for sheet_name,blocks in sorted(blocks.items(), key=lambda e:e[0]) :
            ws = wb.create_sheet(title=sheet_name)
            _c = 0
            
            #------------------------------------------------------------------
            for b,block in enumerate(blocks) :
                lines = []
                
                if len(block['dims']) == 1 :
                    lines.append(block['axes'] + block['variables'])
                    NR = len(block['values'][block['axes'][0]])
                    for r in range(NR) :
                        line = []
                        for c,ax in enumerate(block['axes']) : line.append(block['values'][ax][r])
                        for c,var in enumerate(block['variables']) : line.append(block['values'][var][r])
                        lines.append(line)
                    
                #--------------------------------------------------------------
                elif len(block['dims']) == 2 :
                    lines = []
                    Xaxes = [ax for ax in block['axes'] if ax.startswith(block['dims'][0])]
                    Yaxes = [ax for ax in block['axes'] if ax.startswith(block['dims'][1])]
                    for xaxe in Xaxes :
                        line = ['']*len(Xaxes)
                        line += list(block['values'][xaxe])
                        lines.append(line)
                        
                    NR = len(block['values'][Yaxes[0]])
                    NC = len(block['values'][Xaxes[0]])
                    for j in range(NR-1,-1,-1) :
                        line = []
                        for yaxe in Yaxes : line.append(block['values'][yaxe][j])
                        for i in range(NC) : line.append(block['values'][block['variable']][j,i])
                        lines.append(line)
                        
                    lines[0][0] = '↓{}, {}→'.format(*block['dims'])
                        
                #--------------------------------------------------------------
                for r,line in enumerate(lines) :
                    for c,field in enumerate(line) :
                        ws.cell(row=r+1, column=_c+c+1, value=field)
                if len(lines) > 0 :
                    _c += len(lines[0])+1
                
        #----------------------------------------------------------------------
        for sheet_name in wb.sheetnames :
            ws = wb[sheet_name]
            for col in ws.columns :
                max_length = 0
                column = col[0].column_letter
                for cell in col : max_length = max([len(str(cell.value)),max_length])
                ws.column_dimensions[column].width = max_length+1
                
        #----------------------------------------------------------------------
        while True :
            try :
                wb.save(filepath)
                break
            except :
                msg = "La sauvegarde des données a échoué :(\nLe fichier '{}' est peut-être déjà ouvert ...".format(os.path.basename(filepath))
                rep = question_box('warning', "Sauvegarde impossible", msg, buttons=[('Yes','Fichier fermé !','check'), ('Cancel','Annuler','close')])
                if rep == 'Cancel' : return
            time.sleep(1)
            
        #----------------------------------------------------------------------
        print("OK")
        
        #----------------------------------------------------------------------
        rep = question_box('question',
                           "Sauvegarde terminée",
                           "La sauvegarde des données est terminée !",
                           buttons=[('Yes','OK','check'),
                                    ('No','Ouvrir','excel'),
                                    ])
        if rep == 'No' :
            os.startfile(filepath)
        
    #==========================================================================
    def add_page(self, page=None, gc=None) :
        #----------------------------------------------------------------------
        if gc is not None : gcs = [gc]
        else : gcs = [Graph_Config('1D', ['X'])]
        
        #----------------------------------------------------------------------
        if page is None :
            page = Page(self, None, Page_Config([[1]]), gcs)
        
        #----------------------------------------------------------------------
        page.place_graphs()
        
        #----------------------------------------------------------------------
        self.pages.append(page)
        self.tab_widget.addTab(page, page.name)
        self.tab_widget.setCurrentWidget(page)
        
        #----------------------------------------------------------------------
        return page
        
    #==========================================================================
    def remove_page(self, page_index=None, tab_index=None) :
        #----------------------------------------------------------------------
        page = self.get_page(page_index=page_index, tab_index=tab_index)
        
        #----------------------------------------------------------------------
        for graph in page.graphs :
            graph.remove_graph()
        
        #----------------------------------------------------------------------
        self.tab_widget.removeTab(self.get_index(page))
        self.pages.remove(page)
        page.deleteLater()
        
    #==========================================================================
    def on_close_other_pages(self, keep=None) :
        #----------------------------------------------------------------------
        N_max = 0
        if keep :
            index = self.get_index(keep)
            self.tab_widget.tab_bar.moveTab(index, 0)
            N_max += 1
        
        #----------------------------------------------------------------------
        while self.tab_widget.count() > N_max :
            self.remove_page(tab_index=N_max)
        
    #==========================================================================
    def rename_page(self, page) :
        #----------------------------------------------------------------------
        self.tab_widget.activage_tab_renaming(self.get_index(page))
        
    #==========================================================================
    def on_file_removed(self, file) :
        #----------------------------------------------------------------------
        for page in self.pages :
            for graph in page.graphs :
                changed = False
                for ge in reversed(graph.graph_elements) :
                    if ge.element.file.index == file.index :
                        graph.remove_element(ge, update=False)
                        changed = True
                if changed :
                    graph.update_graph()
    
    #==========================================================================
    def inspector_toggled(self) :
        #----------------------------------------------------------------------
        for page in self.pages :
            for graph in page.graphs :
                graph.update_pointed()
        
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
class Page_Dialog(QDialog) :

    #==========================================================================
    def __init__(self, pm) :
        #----------------------------------------------------------------------
        self.pm = pm
        
        #----------------------------------------------------------------------
        QDialog.__init__(self)

        #----------------------------------------------------------------------
        self.graph_combos = None
        
        #----------------------------------------------------------------------
        self.load_ui()
        self.hide()

    #==========================================================================
    def load_ui(self) :
        #----------------------------------------------------------------------
        uic.loadUi(os.path.join(os.path.dirname(os.path.abspath(__file__)), "UI", "Page_Dialog.ui"), self)
        self.setWindowIcon(self.pm.mw.windowIcon())
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        
        #----------------------------------------------------------------------
        N = list(np.unique([config.N for config in self.pm.page_configs]))
        self.combo_number.clear()
        for n in N :
            self.combo_number.addItem('{} graphique'.format(n)+('s' if n > 1 else ''))
            self.combo_number.setItemData(self.combo_number.count()-1, n)

        #----------------------------------------------------------------------
        self.combo_number.currentIndexChanged.connect(self.on_number_changed)
        self.combo_config.currentIndexChanged.connect(self.on_disposition_changed)
        self.but_valid.clicked.connect(self.on_valid)
        self.but_cancel.clicked.connect(self.close)
        
        #----------------------------------------------------------------------
        self.on_number_changed(self.combo_number.currentIndex())
        self.configure()
        
    #==========================================================================
    def configure(self) :
        #----------------------------------------------------------------------
        self.le_name.setText("Page_{}".format(Page.INDEX))
        self.le_name.setPlaceholderText("Page_{}".format(Page.INDEX))
        self.combo_number.setCurrentIndex(0)
        
    #==========================================================================
    def on_number_changed(self, n) :
        #----------------------------------------------------------------------
        self.combo_config.blockSignals(True)

        #----------------------------------------------------------------------
        self.combo_config.clear()
        for config in self.pm.page_configs :
            if config.N != n+1 : continue
            self.combo_config.addItem(config.icon, config.name)
            self.combo_config.setItemData(self.combo_config.count()-1, config)

        #----------------------------------------------------------------------
        self.combo_config.blockSignals(False)
        self.on_disposition_changed()

    #==========================================================================
    def on_disposition_changed(self) :
        #----------------------------------------------------------------------
        clear_layout(self.layout_graphs)

        #----------------------------------------------------------------------
        page_config = self.combo_config.itemData(self.combo_config.currentIndex())
        self.graph_combos = []
        for e,(r0,c0,r1,c1) in enumerate(page_config.grid_indices) :
            combo = QComboBox()
            combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            combo.setMaxVisibleItems(len(self.pm.graph_configs))
            for graph_config in self.pm.graph_configs :
                combo.addItem(graph_config.name, graph_config.get_icon())
                combo.setItemData(combo.count()-1, graph_config)
            self.graph_combos.append(combo)
            self.layout_graphs.addWidget(combo, r0, c0, r1-r0+1, c1-c0+1)

        #----------------------------------------------------------------------
        self.layout_graphs.update()

    #==========================================================================
    def on_valid(self) :
        #----------------------------------------------------------------------
        page_config = self.combo_config.itemData(self.combo_config.currentIndex())
        graph_configs = [combo.itemData(combo.currentIndex()) for combo in self.graph_combos]

        #----------------------------------------------------------------------
        name = self.le_name.text().strip()
        if name == '' : name = self.le_name.placeholderText()

        #----------------------------------------------------------------------
        new_page = Page(self.pm, name, page_config, graph_configs)
        self.pm.add_page(new_page)
        
        #----------------------------------------------------------------------
        self.hide()

        #----------------------------------------------------------------------
        return new_page

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
class Pages_TabWidget(QTabWidget) :

    #==========================================================================
    page_added   = pyqtSignal(int)
    page_removed = pyqtSignal(int)
    page_renamed = pyqtSignal(int)
    
    #==========================================================================
    def __init__(self, pm) :
        #----------------------------------------------------------------------
        self.pm = pm
        self.mw = self.pm.mw
        
        #----------------------------------------------------------------------
        QTabWidget.__init__(self)
        
        #----------------------------------------------------------------------
        self.tab_bar = Pages_TabBar(self)
        self.setTabBar(self.tab_bar)
        
        #----------------------------------------------------------------------
        self.tabCloseRequested.connect(lambda tab_index : self.pm.remove_page(tab_index=tab_index))
        
        #----------------------------------------------------------------------
        self.le_rename = Pages_LineEdit(self)
        
        self.le_rename.editingFinished.connect(self.on_tab_renamed)
        self.mw.click.connect(self.le_rename.hide)
        self.le_rename.hide()
        
    #==========================================================================
    def mouseReleaseEvent(self, event) :
        #----------------------------------------------------------------------
        if event.button() == 2 :
            self.open_menu()
        
    #==========================================================================
    def open_menu(self) :
        #----------------------------------------------------------------------
        menu = QMenu(self.mw)
        
        #----------------------------------------------------------------------
        menu_add = QMenu("Nouveau graphe", self.mw)
        menu_add.setIcon(get_icon('add'))
        Graph_Config.add_to_menu(menu_add, pm=self.pm, mode='add')
        menu.addMenu(menu_add)
        
        #----------------------------------------------------------------------
        menu.exec_(QCursor.pos())
        
    #==========================================================================
    def addTab(self, page, name) :
        #----------------------------------------------------------------------
        QTabWidget.addTab(self, page, name)
        self.page_added.emit(page.index)
        
    #==========================================================================
    def removeTab(self, index) :
        #----------------------------------------------------------------------
        page = self.pm.get_page(tab_index=index)
        QTabWidget.removeTab(self, index)
        self.page_removed.emit(page.index)
        
    #==========================================================================
    def activage_tab_renaming(self, index) :
        #----------------------------------------------------------------------
        self.renamed_index = index
        page = self.widget(self.renamed_index)
        
        self.le_rename.setText(page.name)
        self.le_rename.setPlaceholderText(page.default_name)
        
        #----------------------------------------------------------------------
        tabRect = self.tab_bar.tabRect(index)
        self.le_rename.setGeometry(QRect(self.tab_bar.mapTo(self, tabRect.topLeft()), tabRect.size()))
        self.le_rename.show()
        self.le_rename.raise_()
        self.le_rename.setFocus()
        self.le_rename.selectAll()
        
    #==========================================================================
    def on_tab_renamed(self) :
        #----------------------------------------------------------------------
        new_name = self.le_rename.text().strip()
        if new_name == '' : new_name = self.le_rename.placeholderText()

        #----------------------------------------------------------------------
        page = self.widget(self.renamed_index)
        page.set_name(new_name)
        
        #----------------------------------------------------------------------
        self.le_rename.hide()
        self.page_renamed.emit(page.index)
        
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
class Pages_TabBar(QTabBar) :

    #==========================================================================
    def __init__(self, tabwidget) :
        #----------------------------------------------------------------------
        self.tabwidget = tabwidget
        self.pm = self.tabwidget.pm
        self.mw = self.pm.mw
        
        #----------------------------------------------------------------------
        QTabBar.__init__(self)
        
        #----------------------------------------------------------------------
        self.setMovable(True)
        self.setTabsClosable(True)
    
    #==========================================================================
    def mousePressEvent(self, event) :
        #----------------------------------------------------------------------
        tab_index = self.tabAt(event.pos())
        self.setCurrentIndex(tab_index)
        
        #----------------------------------------------------------------------
        if tab_index != -1 :
            if event.button() == Qt.MiddleButton :
                self.pm.remove_page(tab_index=tab_index)
                
        #----------------------------------------------------------------------
        QTabBar.mousePressEvent(self, event)
    
    #==========================================================================
    def mouseReleaseEvent(self, event) :
        #----------------------------------------------------------------------
        tab_index = self.tabAt(event.pos())
        if tab_index != -1 :
            if event.button() == Qt.RightButton :
                self.open_menu(event, tab_index)
                
        #----------------------------------------------------------------------
        QTabBar.mouseReleaseEvent(self, event)

    #==========================================================================
    def mouseDoubleClickEvent(self, event) :
        #----------------------------------------------------------------------
        tab_index = self.tabAt(event.pos())
        if tab_index != -1 :
            if event.button() == Qt.LeftButton :
                self.tabwidget.activage_tab_renaming(tab_index)

    #==========================================================================
    def open_menu(self, event, tab_index) :
        #----------------------------------------------------------------------
        menu = QMenu(self.mw)
        #----------------------------------------------------------------------
        page = self.pm.get_page(tab_index=tab_index)
        #----------------------------------------------------------------------
        if len(self.pm.pages) > 1 :
            menu.addAction(get_icon('close')  , "Fermer les autres").triggered.connect(lambda e,page=page : self.pm.on_close_other_pages(keep=page))
            menu.addAction(get_icon('close')  , "Fermer tous").triggered.connect(lambda e : self.pm.on_close_other_pages())
        #----------------------------------------------------------------------
        menu.exec_(event.globalPos())
        
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
class Pages_LineEdit(QLineEdit) :
        
    #==========================================================================
    def keyPressEvent(self, event) :
        #----------------------------------------------------------------------
        if event.key() == Qt.Key_Escape and self.isVisible() :
            self.editingFinished.emit()
            event.accept()
        #----------------------------------------------------------------------
        else :
            super().keyPressEvent(event)
        
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

