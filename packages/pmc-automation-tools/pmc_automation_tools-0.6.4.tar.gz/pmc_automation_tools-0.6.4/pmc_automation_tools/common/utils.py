from datetime import datetime, date, timedelta, timezone
import pytz
import os
import sys
import json
import csv
from warnings import warn
import logging
from logging.handlers import MemoryHandler

from typing import Union, Generator, List, Dict, Any
from openpyxl import load_workbook

DEFAULT_FORMATTER = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
STDOUT_FORMATTER = "[%(asctime)s][%(filename)s:%(lineno)s][%(funcName)20s()] %(message)s"
LOG_FORMATS = {
    "DAILY": "%Y_%m_%d_",
    "MONTHLY": "%Y_%m_"
    }

def debug_logger(level=logging.NOTSET):
    logger = logging.getLogger(__name__)
    logging.basicConfig(format=STDOUT_FORMATTER)
    logger.setLevel(level)
    return logger


def frozen_check():
    """
    Checks the running script to see if it is compiled to a single exe.
    If compiled, the resources will be stored in a temp folder.
    If not, then they will be in the script's working directory.
    """
    if getattr(sys, 'frozen', False):
    # Running in a bundle
        bundle_dir = sys._MEIPASS # pylint: disable=no-member
    else:
    # Running in a normal Python environment
        bundle_dir = os.path.dirname(os.path.abspath(__file__))
    return bundle_dir


def debug_dump_variables(obj):
    """
    Dumps variables of provided object to a log file.
    """
    if not hasattr(obj, 'dump_logger'):
        if hasattr(obj, 'batch_folder'):
            root = obj.batch_folder
        else:
            root = os.getcwd()
        obj.dump_logger = obj.setup_logger('Debug Dump', log_file='Debug_Dump', root_dir=root)
    obj.dump_logger.debug(f"Dumping variables for {type(obj.__name__)}:")
    for k, v in vars(obj).items():
        obj.debug_logger.debug(k, v)
        obj.dump_logger.debug(f'{k} : {v}')


def get_case_insensitive_key_value(input_dict, key):
    return next((value for dict_key, value in input_dict.items() if dict_key.lower() == key.lower()), None)


def create_batch_folder(root:str='', batch_code:str=None, include_time:bool=False, test:bool=False, batch_prefix:bool=None) -> str:
    """
    Used to set up a batch folder to store any log files or screenshots during an automation run.
    
    Parameters:

    - root: The root directory where the batch folder will be created.
    - batch_code: Optional batch code to use instead of generated one. Overrides include_time parameter.
    - include_time: If True, appends the current time to the batch code.
    - test: If True, uses 'TEST' for the batch folder path; otherwise, uses 'PROD'.
    - batch_prefix: adds prefix value to batch code.
    
    Returns:
    
    - The path to the created batch folder.
    """
    if batch_code and include_time:
        warn('batch_code and time arguments are not supported together. Ignoring time argument.')
        include_time = False

    db = 'TEST' if test else 'PROD'
    now = datetime.now()
    b_code = batch_code or now.strftime('%Y%m%d')
    b_time = now.strftime('%H%M')
    folder_name = f'{b_code}_{b_time}' if include_time else b_code
    folder_name = f'{batch_prefix}_{folder_name}' if batch_prefix else folder_name
    batch_folder = os.path.join(root, 'batch_codes', db, folder_name)
    os.makedirs(batch_folder, exist_ok=True)
    return batch_folder


def setup_logger(name:str,
                 log_file:str='log.log',
                 file_format:str='DAILY',
                 level:Union[int, str]=logging.DEBUG,
                 formatter:str=DEFAULT_FORMATTER,
                 root_dir:str=None,
                 flush_level:Union[int, str]=logging.ERROR,
                 write_stdout:bool=True,
                 propagate:bool=False) -> logging.Logger:
    """
    Setup a logger with a memory buffer that flushes on errors, and optionally outputs to stdout.

    Parameters:
    - name: logger name
    - log_file: filename for the log file.
    - file_format: "DAILY" | "MONTHLY" | "". Will be combined with the log_file filename provided.
    - level: log level for the logger. logging module levels.
    - formatter: logging formatter.
    - root_dir: root directory to store the log file.
    - flush_level: log level to trigger flushing to the file.
    - write_stdout: if True, logs will also be printed to stdout.
    - propagate: Do not propogate the logger. Default False since the debug_logger method breaks if it is a parent logger.

    Default formatter: %(asctime)s - %(name)s - %(levelname)s - %(message)s
    """
    date_format = LOG_FORMATS.get(file_format.upper(), "")
    log_date = datetime.now().strftime(date_format) if date_format else ""
    log_path = os.path.join(root_dir, log_date + log_file) if root_dir else log_date + log_file

    file_handler = logging.FileHandler(log_path, mode='a', encoding='utf-8')
    file_handler.setFormatter(logging.Formatter(formatter))

    memory_handler = MemoryHandler(
        capacity=1024,
        flushLevel=flush_level,
        target=file_handler
    )
    
    logger = logging.getLogger(name)
    logger.propagate = propagate
    logger.setLevel(level)
    
    if not logger.hasHandlers():
        logger.addHandler(memory_handler)
        if write_stdout:
            stream_handler = logging.StreamHandler()
            stream_handler.setFormatter(logging.Formatter(STDOUT_FORMATTER))
            stream_handler.setLevel(level)
            logger.addHandler(stream_handler)
    return logger

def read_updated(in_file:str, obj_type:Union[dict, list]=None, sheet_name:str=None) -> Union[List[Dict[str, str]],Any]: 
    
    """
    Read in a file of already updated records.

    Parameters:
    
    - in_file: file containing the data to read.
    - obj_type: default object type to return if file is empty or doesn't exist.

    Returns:

    - json object from file or empty obj_type object
    - list of dictionaries containing csv row data.
    """
    if obj_type is None:
        obj_type = []
    updated_records = obj_type
    _file_type = in_file.split('.')[-1].lower()
    if _file_type in ['xlsx', 'xlsm']:
        return _read_excel(in_file, sheet_name)
    if os.path.exists(in_file) and os.path.getsize(in_file) > 0:
        with open(in_file, 'r', encoding='utf-8-sig') as f:
            if _file_type == 'json':
                updated_records = json.load(f)
            elif _file_type == 'csv':
                c = csv.DictReader(f)
                updated_records = [row for row in c]
            else:
                raise TypeError('File name provided is not an expected type of xlsx, json or csv.')
    return updated_records


def _read_excel(file_path:str, sheet_name:str=None) -> List[Dict[str, str]]:
    """
    Reads the contents of the first sheet in an Excel (.xlsx) file.

    Parameters:
    - file_path: Path to the Excel file to read.

    Returns:
    - List of dictionaries, where each dictionary represents a row, with keys as column headers.
    """
    while True:
        try:
            workbook = load_workbook(file_path)
            break
        except PermissionError:
            print("\a")
            user_input = input("\n[WARNING] The file is currently in use. Please close the file and press Enter to try again, or type 'cancel' to stop: ")
            if user_input.lower() == 'cancel':
                return []
    sheets = workbook.sheetnames
    if sheet_name is None:
        first_sheet = workbook[sheets[0]]
    elif sheet_name in sheets:
        first_sheet = workbook[sheet_name]
    else:
        raise ValueError(f"Sheet name {sheet_name} not found in workbook.")
    data = []
    
    # Get the headers (first row)
    headers = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    
    # Iterate over the remaining rows and build dictionaries
    for row in first_sheet.iter_rows(min_row=2, values_only=True):
        row_strings = [str(cell) if cell is not None else "" for cell in row]
        row_data = dict(zip(headers, row_strings))
        data.append(row_data)
    
    return data


def save_updated_overwrite(in_file:str, obj:dict) -> None:
    """
    Save a file containing a list of already processed records.

    Parameters:
    
    - in_file: file to use to save
    - obj: json object to write to file. Expects a list containing dictionaries.
    """
    if not obj:
        return
    _file_type = in_file.split('.')[-1].lower()
    with open(in_file, 'w+', encoding='utf-8') as f:
        if _file_type == 'json':
            f.write(json.dumps(obj, indent=4))
        elif _file_type == 'csv':
            c = csv.DictWriter(f, fieldnames=obj[0].keys(), lineterminator='\n')
            c.writeheader()
            c.writerows(obj)
        else:
            raise TypeError('File name provided is not an expected type of json or csv.')
        

def save_updated(in_file:str, obj:Union[dict, list], overwrite:bool=False) -> None:
    """
    Append to a file containing a list of already processed records.

    Parameters:
    
    - in_file: file to use to save
    - obj: json object to write to file.
    - overwrite: use the overwrite version of the update regardless of provided object.
    """
    if not obj:
        return
    if isinstance(obj, list) or overwrite:
        return save_updated_overwrite(in_file, obj)
    obj = [obj]
    _file_type = in_file.split('.')[-1].lower()
    
    # Append JSON data
    if _file_type == 'json':
        # Load existing data if the file exists and is non-empty
        if os.path.exists(in_file) and os.path.getsize(in_file) > 0:
            with open(in_file, 'r', encoding='utf-8') as f:
                try:
                    existing_data = json.load(f)
                except json.JSONDecodeError:
                    existing_data = []
        else:
            existing_data = []
        
        # Append new data and save
        with open(in_file, 'w', encoding='utf-8') as f:
            json.dump(existing_data + obj, f, indent=4)

    # Append CSV data
    elif _file_type == 'csv':
        file_exists = os.path.isfile(in_file)
        with open(in_file, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=obj[0].keys(), lineterminator='\n')
            
            # Write the header only if the file is new or empty
            if not file_exists or os.path.getsize(in_file) == 0:
                writer.writeheader()
            
            writer.writerows(obj)
    else:
        raise TypeError('File name provided is not an expected type of json or csv.')
    

def plex_date_formatter(*args: datetime|int, date_offset:int=0, tz_convert:bool=True, tz:str="America/New_York") -> str:
    """
    Takes 'normal' date formats and converts them to a Plex web service format (ISO format)
    Can also take a single datetime object.
    2022, 09, 11 -> 2022-09-11T04:00:00Z
    2022, 09, 11, 18, 45 -> 2022-09-11T22:45:00Z
        Next day if hours fall into 20-24 period
    2022, 09, 11, 22 -> 2022-09-12T02:00:00Z
        date_offset arg will add days to the provided time
        Useful when providing just a datetime object to the function
    """
    if isinstance(args[0], datetime):
        _date = args[0]
    elif isinstance(args[0], date):
        _date = datetime.combine(args[0], datetime.min.time())
    else:
        _date = datetime(*args)
    if tz_convert:
        _tz = pytz.timezone(tz)
        _date = _tz.localize(_date).astimezone(timezone.utc)
    _date += timedelta(days=date_offset)
    f_date = _date.strftime('%Y-%m-%dT%H:%M:%SZ')
    return f_date


def chunk_list(lst:list, chunk_size:int) -> Generator[list, None, None]:
    for i in range(0, len(lst), chunk_size):
        yield lst[i:i + chunk_size]