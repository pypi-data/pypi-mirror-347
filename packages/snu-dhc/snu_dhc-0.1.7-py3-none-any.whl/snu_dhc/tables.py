#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from scipy.stats import ttest_ind, mannwhitneyu, chi2_contingency, f_oneway, kruskal, normaltest, fisher_exact, levene
from typing import Union, Literal
import warnings

class DemoTable:
    def __init__(self, df: pd.DataFrame, *,
                 table_name: str,
                 variables: list,
                 group_variable: str = None,
                 group_labels: dict = None,
                 show_total: bool = True,
                 show_missing: bool = True,
                 percent_decimals: int = 1,
                 thousands_sep: bool = True,
                 show_p_values: bool = False,
                 p_value_decimals: Union[int, Literal["auto"]] = "auto"):
        """
        Initializes a DemoTable for generating baseline characteristics Excel tables.

        Parameters
        ----------
        df : pd.DataFrame
            Input dataframe containing the raw data.

        table_name : str
            Title to be shown at the top of the Excel sheet.

        variables : list of dict
            List of variables to include in the table.
            Each dict must include:
                - "var" : str  → column name in df
                - "name" : str  → printed variable name
                - "type" : "categorical" or "continuous"
                - For continuous:
                    - "stat" : "mean" or "median"
                    - Optional: "decimals" : int (decimal places for mean/std or median/IQR)
                - For categorical:
                    - "class_labels" : dict {value: label_to_display}

            Example:
            variables = [
                {"var": "age", "name": "Age", "type": "continuous", "stat": "mean", "decimals": 1},
                {"var": "sex", "name": "Sex", "type": "categorical", "class_labels": {1: "Male", 0: "Female"}}
            ]

        group_variable : str, optional
            Column name to use for group comparison (e.g., 'treatment_group').

        group_labels : dict, optional
            Dictionary mapping group values to display names.
            Example: {1: "Treatment", 0: "Control"}

        show_total : bool, default=True
            Whether to include a "Total" column.

        show_missing : bool, default=True
            Whether to display missing values as a separate category in categorical variables.

        percent_decimals : int, default=1
            Number of decimal places to show in percentages.

        thousands_sep : bool, default=True
            Whether to use commas in large numbers (e.g., 1,234).

        show_p_values : bool, default=False
            Whether to include a column for p-values.

            Test selection logic:
            - Continuous variables with `stat='mean'`:
                * 2 groups → Welch's t-test (normality & variance checked, warnings issued if violated)
                * ≥3 groups → ANOVA (with normality and variance checks, warnings issued)
            - Continuous variables with `stat='median'`:
                * 2 groups → Mann–Whitney U test
                * ≥3 groups → Kruskal–Wallis test
            - Categorical variables:
                * Default: chi-square test
                * If 2×2 table and any expected count < 5 → Fisher’s exact test
                * If >2×2 table and any expected count < 5 → fall back to chi-square with warning

        p_value_decimals : int or "auto", default="auto"
            Number of decimal places for p-values.
            - If int: fixed number of decimals.
            - If "auto": 
                * p < 0.001 → "<0.001"
                * p > 0.99 → ">0.99"
                * p < 0.01 → 3 decimals
                * 0.045 ≤ p < 0.055 → 3 decimals (to avoid ambiguity near 0.05)
                * Otherwise → 2 decimals
        """

        self.df = df
        self.table_name = table_name
        self.variables = variables
        self.group_variable = group_variable
        self.group_labels = group_labels or {}
        self.show_total = show_total
        self.show_missing = show_missing
        self.percent_decimals = percent_decimals
        self.thousands_sep = thousands_sep
        self.show_p_values = show_p_values
        self.p_value_decimals = p_value_decimals

        if not isinstance(df, pd.DataFrame):
            raise TypeError("df must be a pandas DataFrame.")

        if not isinstance(variables, list) or not all(isinstance(v, dict) for v in variables):
            raise TypeError("variables must be a list of dictionaries.")

        if not isinstance(percent_decimals, int):
            raise TypeError("percent_decimals must be an integer.")

        if not (isinstance(p_value_decimals, int) or p_value_decimals == "auto"):
            raise ValueError("p_value_decimals must be an integer or 'auto'.")

    def save(self, filepath: str):
        table_df = self._build_table()
        self._write_excel(table_df, filepath)

    def _format_number(self, num, decimals):
        if pd.isnull(num):
            return ""
        fmt = f"{{:,.{decimals}f}}" if self.thousands_sep else f"{{:.{decimals}f}}"
        return fmt.format(num)

    def _categorical_summary(self, series, class_labels):
        counts = series.value_counts(dropna=False)
        total = counts.sum()
        summary = {}

        for cls, label in class_labels.items():
            n = counts.get(cls, 0)
            pct = 100 * n / total if total > 0 else 0
            formatted = f"{n:,} ({pct:.{self.percent_decimals}f}%)" if self.thousands_sep else f"{n} ({pct:.{self.percent_decimals}f}%)"
            summary[label] = formatted

        # Proper handling of missing values
        if self.show_missing:
            missing_count = series.isnull().sum()
            if missing_count > 0:
                pct = 100 * missing_count / total if total > 0 else 0
                formatted = f"{missing_count:,} ({pct:.{self.percent_decimals}f}%)" if self.thousands_sep else f"{missing_count} ({pct:.{self.percent_decimals}f}%)"
                summary["Missing"] = formatted

        return summary

    def _continuous_summary(self, series, stat, decimals):
        series = series.dropna()
        if len(series) == 0:
            return ""
        if stat == 'mean':
            mean = np.mean(series)
            std = np.std(series, ddof=1)
            return f"{self._format_number(mean, decimals)} ({self._format_number(std, decimals)})"
        elif stat == 'median':
            med = np.median(series)
            q1 = np.percentile(series, 25)
            q3 = np.percentile(series, 75)
            return f"{self._format_number(med, decimals)} ({self._format_number(q1, decimals)}–{self._format_number(q3, decimals)})"
        else:
            return ""

    def _format_p(self, p):
        if pd.isnull(p):
            return ""

        if p < 0.001:
            return "<0.001"
        if p > 0.99:
            return ">0.99"

        if isinstance(self.p_value_decimals, int):
            return f"{p:.{self.p_value_decimals}f}"

        # auto mode
        if p < 0.01:
            return f"{p:.3f}"
        elif 0.045 <= p < 0.055:
            return f"{p:.3f}"  # avoid ambiguity near 0.05
        else:
            return f"{p:.2f}"

    def _compute_p_value(self, var_cfg):
        if not self.group_variable:
            return None

        group_vals = [g for g in self.group_labels if g in self.df[self.group_variable].unique()]
        grouped_data = [self.df[self.df[self.group_variable] == g][var_cfg['var']].dropna() for g in group_vals]

        if var_cfg["type"] == "continuous":
            if var_cfg["stat"] == "mean":
                if len(grouped_data) == 2:
                    g1, g2 = grouped_data

                    # Normality
                    norm1 = normaltest(g1)[1] if len(g1) >= 8 else 1.0
                    norm2 = normaltest(g2)[1] if len(g2) >= 8 else 1.0
                    if norm1 < 0.05 or norm2 < 0.05:
                        print(
                            f"[Note] Non-normal distribution detected for variable '{var_cfg['name']}'. "
                            f"Consider using Mann–Whitney U test instead of t-test."
                        )

                    # Variance equality
                    if len(g1) > 1 and len(g2) > 1:
                        lev_p = levene(g1, g2)[1]
                        if lev_p < 0.05:
                            print(
                                f"[Note] Unequal variances detected for variable '{var_cfg['name']}' "
                                f"(Levene’s test p={lev_p:.3f}). Using Welch’s t-test."
                            )

                    _, p = ttest_ind(g1, g2, equal_var=False)

                elif len(grouped_data) > 2:
                    normality_pvals = [normaltest(g)[1] for g in grouped_data if len(g) >= 8]
                    if any(p < 0.05 for p in normality_pvals):
                        print(
                            f"[Note] Non-normal distribution detected in one or more groups for variable '{var_cfg['name']}'. "
                            f"Consider using Kruskal–Wallis test instead of ANOVA."
                        )

                    if all(len(g) > 1 for g in grouped_data):
                        lev_p = levene(*grouped_data)[1]
                        if lev_p < 0.05:
                            print(
                                f"[Note] Unequal variances detected among groups for variable '{var_cfg['name']}' "
                                f"(Levene’s test p={lev_p:.3f}). ANOVA may not be appropriate."
                            )

                    _, p = f_oneway(*grouped_data)
                else:
                    p = None

            elif var_cfg["stat"] == "median":
                if len(grouped_data) == 2:
                    _, p = mannwhitneyu(grouped_data[0], grouped_data[1], alternative='two-sided')
                elif len(grouped_data) > 2:
                    _, p = kruskal(*grouped_data)
                else:
                    p = None
            else:
                p = None

        elif var_cfg["type"] == "categorical":
            contingency = pd.crosstab(self.df[var_cfg['var']], self.df[self.group_variable])
            if contingency.shape[0] > 1 and contingency.shape[1] > 1:
                if (contingency < 5).any().any():
                    if contingency.shape == (2, 2):
                        print(
                            f"[Note] Low expected cell count detected for variable '{var_cfg['name']}'. "
                            f"Using Fisher's exact test instead of chi-square."
                        )
                        try:
                            _, p = fisher_exact(contingency.values)
                        except:
                            p = None
                    else:
                        print(
                            f"[Note] Low expected cell count detected for variable '{var_cfg['name']}', "
                            f"but Fisher's exact test only supports 2x2 tables. Falling back to chi-square test."
                        )
                        try:
                            _, p, _, _ = chi2_contingency(contingency)
                        except:
                            p = None
                else:
                    try:
                        _, p, _, _ = chi2_contingency(contingency)
                    except:
                        p = None
            else:
                p = None
        else:
            p = None

        return p

    def _build_table(self):
        rows = []
        groups = []

        # Determine groups
        if self.group_variable:
            groups = [g for g in self.group_labels if g in self.df[self.group_variable].unique()]
            if self.show_total:
                groups = [None] + groups
        else:
            groups = [None]

        # Disable p-values if group info is missing
        show_p = self.show_p_values and self.group_variable is not None and self.group_labels

        # Build column headers
        colnames = ["Variable", "Class"] + ([self.group_labels.get(g, str(g)) if g is not None else "Total" for g in groups])
        if show_p:
            colnames.append("P-value")

        # Row for counts
        count_row = ["N", None]
        for g in groups:
            subset = self.df if g is None else self.df[self.df[self.group_variable] == g]
            n = len(subset)
            count_row.append(f"{n:,}" if self.thousands_sep else str(n))
        if show_p:
            count_row.append("")
        rows.append(count_row)

        # Variable rows
        for var_cfg in self.variables:
            var = var_cfg["var"]
            name = var_cfg["name"]
            typ = var_cfg["type"]
            p_val = self._compute_p_value(var_cfg) if show_p else None
            p_text = self._format_p(p_val) if p_val is not None else ""

            if typ == "categorical":
                class_labels = var_cfg.get("class_labels", {})

                # Aggregate summaries from all groups
                combined_summary = {}
                for g in groups:
                    subset = self.df if g is None else self.df[self.df[self.group_variable] == g]
                    summary = self._categorical_summary(subset[var], class_labels)
                    for label in summary:
                        if label not in combined_summary:
                            combined_summary[label] = {}
                        combined_summary[label][g] = summary[label]

                # Generate rows
                for i, (label, values_by_group) in enumerate(combined_summary.items()):
                    row = [name if i == 0 else None, label]
                    for g in groups:
                        row.append(values_by_group.get(g, "0 (0.0%)"))
                    if show_p:
                        row.append(p_text if i == 0 else "")
                    rows.append(row)

            elif typ == "continuous":
                stat = var_cfg["stat"]
                decimals = var_cfg.get("decimals", self.percent_decimals)
                row = [name, None]
                for g in groups:
                    subset = self.df if g is None else self.df[self.df[self.group_variable] == g]
                    summary = self._continuous_summary(subset[var], stat, decimals)
                    row.append(summary)
                if show_p:
                    row.append(p_text)
                rows.append(row)

        return pd.DataFrame(rows, columns=colnames)

    def _write_excel(self, table_df: pd.DataFrame, filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Demographics"

        thin_border = Border(bottom=Side(style='thin', color='000000'))

        # Title row (row 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table_df.shape[1])
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = self.table_name
        title_cell.font = Font(bold=True, size=11, name="Times New Roman")
        title_cell.alignment = Alignment(horizontal='left')

        for c in range(1, table_df.shape[1] + 1):
            ws.cell(row=1, column=c).border = thin_border

        # Write data rows starting at row 2
        max_col_widths = [len(str(col)) for col in table_df.columns]  # start with header widths

        for r_idx, row in enumerate(dataframe_to_rows(table_df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.font = Font(name="Times New Roman")
                cell.alignment = Alignment(horizontal='left')

                # Update max column width if needed
                val_len = len(str(value)) if value is not None else 0
                if val_len > max_col_widths[c_idx - 1]:
                    max_col_widths[c_idx - 1] = val_len

            # Apply border to header row
            if r_idx == 2:
                for c_idx in range(1, table_df.shape[1] + 1):
                    ws.cell(row=r_idx, column=c_idx).border = thin_border

        # Apply border to last data row
        for c_idx in range(1, table_df.shape[1] + 1):
            ws.cell(row=r_idx, column=c_idx).border = thin_border

        # Adjust column widths
        for i, width in enumerate(max_col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width + 1

        wb.save(filepath)
        print(f"Saved Excel file to: {filepath}")

