# dh_tool/dataframe/handlers/excel_handler.py
import json
import numpy as np
import pandas as pd
from datetime import datetime
import warnings

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from ..utils.events import EventEmitter


class ExcelHandler(EventEmitter):
    def __init__(self):
        super().__init__()
        self.df = None
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Sheet1"  # 기본 시트 이름 설정
        self.sheet_names = {"Sheet1"}
        self.add_column_prefix = False  # 접두사 추가 여부 설정

    def update(self, dataframe):
        self.df = dataframe
        self.sync_df_to_worksheet()

    def sync_df_to_worksheet(self):
        self.worksheet.delete_rows(1, self.worksheet.max_row)

        if self.add_column_prefix:
            # 열 이름에 접두사 추가 (이미 'col_'로 시작하는 경우 제외)
            self.df.columns = [
                col if col.startswith("col_") else f"col_{col}"
                for col in self.df.columns
            ]

        for row in dataframe_to_rows(self.df, index=False, header=True):
            row = [self._convert_to_string_if_needed(cell) for cell in row]
            self.worksheet.append(row)

    def _convert_to_string_if_needed(self, value):
        """복잡한 데이터 타입을 문자열로 변환"""
        try:
            if isinstance(value, (dict, list, np.ndarray)):
                warnings.warn(
                    f"Complex data type {type(value)} will be converted to string representation"
                )
                return str(value)  # json.dumps 대신 str 사용
            elif pd.isna(value):
                return ""
            elif isinstance(value, (np.int64, np.float64)):
                return float(value)
            elif isinstance(value, datetime):
                return value.isoformat()
            elif value in (np.inf, -np.inf):
                return str(value)
            return value
        except Exception as e:
            warnings.warn(
                f"Error converting value {value}: {str(e)}. Using string representation."
            )
            return str(value)

    def set_column_width(self, **kwargs):
        """컬럼 너비 설정"""
        for column, width in kwargs.items():
            if column in self.df.columns:
                col_idx = self.df.columns.get_loc(column) + 1
                col_letter = chr(64 + col_idx)
                self.worksheet.column_dimensions[col_letter].width = width
            else:
                pass

    def freeze_first_row(self):
        """첫 번째 행 고정"""
        self.worksheet.freeze_panes = self.worksheet["A2"]

    def _apply_autowrap(self):
        """자동 줄바꿈 스타일 적용"""
        for row in self.worksheet.iter_rows(min_row=1, max_row=self.worksheet.max_row):
            for cell in row:
                if cell.value:  # 값이 있는 셀에만 적용
                    cell.alignment = Alignment(
                        wrap_text=True, vertical="center"
                    )  # 수직 정렬도 추가

    def enable_autowrap(self):
        """자동 줄바꿈 활성화"""
        self._apply_autowrap()

    def add_hyperlink(self, cell, url, display=None):
        """셀에 하이퍼링크 추가"""
        self.worksheet[cell].hyperlink = url
        self.worksheet[cell].value = display if display else url
        self.worksheet[cell].style = "Hyperlink"

    def add_hyperlinks_to_column(self, column_name, urls, display_texts=None):
        """컬럼에 있는 각 셀에 하이퍼링크 추가"""
        if len(urls) != len(self.df):
            raise ValueError(
                "The length of the URL list must match the length of the dataframe."
            )

        if display_texts and len(display_texts) != len(self.df):
            raise ValueError(
                "The length of the display_texts list must match the length of the dataframe."
            )

        col_idx = self.df.columns.get_loc(column_name) + 1
        for i, url in enumerate(urls, start=2):  # start=2 to account for header row
            cell = f"{chr(64 + col_idx)}{i}"
            display = display_texts[i - 2] if display_texts else None
            self.add_hyperlink(cell, url, display)

    def save(self, filename):
        """엑셀 파일 저장"""
        self.workbook.save(filename)

    def close(self):
        """엑셀 파일 닫기"""
        del self.workbook

    def __del__(self):
        self.close()

    # 시트 관련 기능 추가
    def create_sheet(self, dataframe, title=None):
        """새 시트 생성 및 데이터프레임 추가"""
        if title is None:
            title = f"Sheet{len(self.sheet_names) + 1}"
        if title in self.sheet_names:
            print(f"{title} sheet은 이미 존재합니다, 이름을 바꿔주세요")
            return False
        self.sheet_names.add(title)
        self.worksheet = self.workbook.create_sheet(title=title)
        self.df_to_worksheet(dataframe)
        self.df = dataframe
        return True

    def select_sheet(self, title):
        """시트 선택"""
        self.worksheet = self.workbook[title]
        data = list(self.worksheet.values)
        cols = data[0]  # First row as columns
        self.df = pd.DataFrame(data[1:], columns=cols)

        # 열 이름에서 접두사 제거 (모든 'col_' 접두사 제거)
        self.df.columns = [col.replace("col_", "") for col in self.df.columns]

        # 데이터 타입 복원
        for col in self.df.columns:
            if self.df[col].dtype == "object":
                try:
                    # JSON 형식의 문자열을 파이썬 객체로 변환 시도
                    self.df[col] = self.df[col].apply(json.loads)
                except:
                    pass
                try:
                    # 날짜 형식 문자열을 datetime 객체로 변환 시도
                    self.df[col] = pd.to_datetime(self.df[col])
                except:
                    pass

                # inf, -inf 처리
                self.df[col] = self.df[col].replace({"inf": np.inf, "-inf": -np.inf})

        return self.df

    def remove_sheet(self, title):
        """시트 제거"""
        sheet_to_remove = self.workbook[title]
        self.workbook.remove(sheet_to_remove)

    def df_to_worksheet(self, dataframe):
        """데이터프레임을 현재 워크시트에 추가"""
        # 데이터프레임을 rows로 변환
        rows = dataframe_to_rows(dataframe, index=False, header=True)

        # 각 행의 각 셀에 대해 데이터 타입 변환 적용
        for row in rows:
            converted_row = [self._convert_to_string_if_needed(cell) for cell in row]
            self.worksheet.append(converted_row)

    def list_sheets(self):
        """시트 목록 반환"""
        return self.workbook.sheetnames

    def get_active_sheet(self):
        """현재 선택된 시트 이름 반환"""
        return self.worksheet.title

    def apply_style(self, style_func):
        """사용자 정의 스타일 함수 적용"""
        for row in self.worksheet.iter_rows():
            for cell in row:
                style_func(cell)

    def color_cells(self, condition, color):
        """조건에 맞는 셀에 색상 적용"""
        for row in self.worksheet.iter_rows():
            for cell in row:
                if condition(cell.value):
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

    def auto_adjust_columns(self):
        """열 너비 자동 조정"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.worksheet.column_dimensions[column_letter].width = adjusted_width

    def add_conditional_formatting(self, range_string, formula, fill):
        """조건부 서식 추가"""
        self.worksheet.conditional_formatting.add(
            range_string, FormulaRule(formula=[formula], fill=fill)
        )
