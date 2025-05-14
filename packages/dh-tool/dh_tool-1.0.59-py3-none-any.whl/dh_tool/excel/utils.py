from openpyxl.utils import get_column_letter
import pandas as pd


import string
import colorsys


def hex_to_argb(hex_color):
    """
    HEX 색상 (#RRGGBB) → ARGB (FFRRGGBB) 변환
    """
    hex_color = hex_color.lstrip("#")  # '#' 제거
    if len(hex_color) != 6:
        raise ValueError("HEX 색상 코드는 6자리여야 합니다 (#RRGGBB).")

    return f"09{hex_color.upper()}"  # 불투명 Alpha(FF) 추가


def generate_color_variants(argb_color, steps=10, lightness_factor=0.1):
    """
    주어진 ARGB 색상(AARRGGBB)에서 명도를 조정하여 steps 단계의 색상 리스트 생성
    - steps: 색상 단계 개수
    - lightness_factor: 명도 변화 정도 (0.1 ~ 0.3 추천)
    - 결과: "AARRGGBB" 형식의 리스트 반환
    """
    if len(argb_color) != 8:
        raise ValueError("ARGB 색상 코드는 8자리(AARRGGBB)여야 합니다.")

    alpha = argb_color[:2]  # Alpha 값 유지
    r, g, b = tuple(int(argb_color[i : i + 2], 16) / 255.0 for i in (2, 4, 6))
    h, l, s = colorsys.rgb_to_hls(r, g, b)  # RGB → HLS 변환

    # 명도 범위 설정 (너무 밝거나 어둡지 않게 조정)
    min_lightness = max(0.2, l - lightness_factor)  # 기존 l보다 어둡게
    max_lightness = min(0.95, l + lightness_factor)  # 기존 l보다 밝게

    variants = []
    for i in range(steps):
        new_l = min_lightness + (max_lightness - min_lightness) * (
            i / (steps - 1)
        )  # 선형 보간
        new_r, new_g, new_b = colorsys.hls_to_rgb(h, new_l, s)  # HLS → RGB 변환
        new_argb = f"{alpha}{int(new_r*255):02X}{int(new_g*255):02X}{int(new_b*255):02X}"  # Alpha 유지
        variants.append(new_argb)

    return variants


# def generate_color_variants(hex_color, steps=10, lightness_factor=0.1):
#     """
#     주어진 HEX 색상에서 명도를 조정하여 steps 단계의 색상 리스트 생성
#     - lightness_factor: 명도 변화의 정도 (0.1 ~ 0.3 추천)
#     """
#     hex_color = hex_color.lstrip("#")  # '#' 제거
#     r, g, b = tuple(
#         int(hex_color[i : i + 2], 16) / 255.0 for i in (0, 2, 4)
#     )  # 0~1 범위 변환
#     h, l, s = colorsys.rgb_to_hls(r, g, b)  # RGB → HSL 변환

#     # 💡 기존 명도를 중심으로 약간씩 조정
#     min_lightness = max(0.2, l - lightness_factor)  # 기존 l보다 살짝 어둡게
#     max_lightness = min(0.95, l + lightness_factor)  # 기존 l보다 살짝 밝게

#     variants = []
#     for i in range(steps):
#         new_l = min_lightness + (max_lightness - min_lightness) * (
#             i / (steps - 1)
#         )  # 선형 보간
#         new_r, new_g, new_b = colorsys.hls_to_rgb(h, new_l, s)  # HSL → RGB 변환
#         new_hex = f"FF{int(new_r*255):02X}{int(new_g*255):02X}{int(new_b*255):02X}"  # ARGB 변환
#         variants.append(new_hex)

#     return variants


def map_column_names_to_letters(worksheet, width_map):
    """
    DataFrame의 컬럼 이름 또는 엑셀 열 문자(A, B, C)를 자동 매핑하여 열 너비 설정
    """
    # 엑셀 시트의 헤더 가져오기
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    # 최종 매핑 결과 저장
    col_letter_map = {}

    for key, width in width_map.items():
        if key in headers:
            # ✅ 컬럼 이름을 엑셀 열 문자로 변환
            col_idx = headers.index(key) + 1
            col_letter = get_column_letter(col_idx)
            col_letter_map[col_letter] = width
        else:
            print(f"컬럼 '{key}'을 찾을 수 없습니다.")

    return col_letter_map


# def get_column_indices_by_condition(worksheet, condition):
#     """
#     조건을 만족하는 열의 인덱스를 반환
#     - condition: 각 열의 데이터 리스트를 받아 True/False를 반환하는 함수
#     """
#     headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
#     indices = []

#     for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
#         col_data = [cell.value for cell in col]
#         if condition(col_data):
#             indices.append(idx)

#     return indices


def get_cell_addresses(df, condition):
    """
    ✅ DataFrame과 조건을 받아 조건을 만족하는 셀의 주소(A1, B2 등)를 반환
    - df: pandas DataFrame
    - condition: 불리언 Series 또는 DataFrame
    """
    cells = []

    # ✅ 1. Series인 경우 (특정 컬럼에만 조건 적용)
    if isinstance(condition, pd.Series):
        col_name = condition.name
        col_idx = df.columns.get_loc(col_name)
        col_letter = get_column_letter(col_idx + 1)  # A, B, ..., Z, AA, AB 지원

        # 조건이 True인 경우 해당 셀 주소 저장
        for row_offset, (row_idx, match) in enumerate(condition.fillna(False).items()):
            if match:
                excel_row = row_offset + 2  # 헤더가 1행, 데이터는 2행부터 시작
                cell_ref = f"{col_letter}{excel_row}"
                cells.append(cell_ref)

    # ✅ 2. DataFrame인 경우 (여러 컬럼에 조건 적용)
    elif isinstance(condition, pd.DataFrame):
        for row_offset, (row_idx, row) in enumerate(condition.fillna(False).iterrows()):
            for col_idx, match in enumerate(row):
                if match:
                    col_letter = get_column_letter(col_idx + 1)
                    excel_row = row_offset + 2
                    cell_ref = f"{col_letter}{excel_row}"
                    cells.append(cell_ref)

    else:
        raise ValueError("Condition must be a Series or DataFrame")

    return cells


# def find_columns_with_nulls(worksheet):
#     """
#     결측치가 있는 컬럼 반환
#     """
#     headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
#     null_columns = []

#     for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
#         col_data = [cell.value for cell in col]
#         if any(pd.isnull(value) for value in col_data):
#             null_columns.append(headers[idx - 1])

#     return null_columns


# def find_columns_by_type(worksheet, data_type):
#     """
#     특정 데이터 타입(int, str 등)을 가진 열 찾기
#     """
#     headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
#     type_columns = []

#     for idx, col in enumerate(worksheet.iter_cols(min_row=2), start=1):
#         col_data = [cell.value for cell in col]
#         if all(isinstance(value, data_type) or pd.isnull(value) for value in col_data):
#             type_columns.append(headers[idx - 1])

#     return type_columns


def apply_to_cells(style_func):
    """
    ✅ 셀 스타일 적용을 제어하는 데코레이터
    - cells=None → 전체 워크시트 적용
    - cells=[...] → 특정 셀 리스트 적용
    - cells=["A1:H1", "B3:D3"] → 범위 지원
    """

    def wrapper(worksheet, *args, cells=None, **kwargs):
        if not cells:  # ✅ 전체 워크시트 적용
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value:
                        style_func(cell, *args, **kwargs)

        else:
            for cell_ref in cells:
                if ":" in cell_ref:  # ✅ 범위 지원 (예: "A1:H1")
                    for row in worksheet[
                        cell_ref
                    ]:  # `range()`를 사용해 범위 내 모든 셀 선택
                        for cell in row:
                            style_func(cell, *args, **kwargs)
                else:  # ✅ 개별 셀 리스트 처리 (예: ["A1", "B1", "C1"])
                    cell = worksheet[cell_ref]
                    style_func(cell, *args, **kwargs)

        return worksheet

    return wrapper


# def get_full_column_ranges(cells, total_rows):
#     """
#     셀 리스트에서 컬럼(A, B, C 등)만 추출하여 전체 컬럼 범위 리스트 반환
#     """
#     columns = {cell[:1] for cell in cells}  # 컬럼만 추출 (A, B, C, ...)
#     return [f"{col}2:{col}{total_rows+1}" for col in columns]  # "A2:A100" 형식 반환


def get_full_row_cells(cells, column_names, target_columns=None):
    """
    특정 셀 리스트에서 해당 행 전체(전체 컬럼 or 특정 컬럼) 범위 반환
    - column_names: 전체 컬럼 이름 리스트 (e.g., df.columns.tolist())
    - target_columns: 특정 컬럼 리스트를 지정하면 해당 컬럼만 적용 (기본값: 전체 컬럼)
    """
    rows = {int(cell[1:]) for cell in cells}  # ✅ 행 번호 추출 (2, 3, 4...)

    # 컬럼명 → 엑셀 컬럼(A, B, C...) 변환
    column_map = {
        name: string.ascii_uppercase[i] for i, name in enumerate(column_names)
    }

    if target_columns is None:
        target_columns = list(column_map.values())  # 기본값: 전체 컬럼 사용
    else:
        target_columns = [
            column_map[col] for col in target_columns
        ]  # 선택된 컬럼만 변환

    return [f"{col}{row}" for row in rows for col in target_columns]
