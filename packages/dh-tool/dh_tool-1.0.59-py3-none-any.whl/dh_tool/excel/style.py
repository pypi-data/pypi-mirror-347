from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from .utils import map_column_names_to_letters, apply_to_cells
from .config import COLOR_MAP


class Style:

    @staticmethod
    def apply_auto_wrap(worksheet):
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

    @staticmethod
    def freeze_panes(worksheet, freeze_pane):
        """엑셀 창 고정"""
        if freeze_pane:
            worksheet.freeze_panes = freeze_pane

    @staticmethod
    def set_column_width(worksheet, width_map):
        """컬럼 이름 또는 엑셀 열 문자로 열 너비 설정"""
        col_letter_map = map_column_names_to_letters(worksheet, width_map)
        for col, width in col_letter_map.items():
            worksheet.column_dimensions[col].width = width

    @staticmethod
    def auto_adjust_column_widths(worksheet):
        """데이터에 맞게 자동으로 열 너비 조정"""
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def apply_auto_filter(worksheet, columns=None):
        """
        데이터 필터 적용
        - columns: None이면 모든 컬럼에 필터 적용
                   리스트면 특정 컬럼에만 필터 적용
        """
        max_row = worksheet.max_row

        if columns is None:  # ✅ 모든 열에 필터 적용
            max_col = worksheet.max_column
            worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        elif isinstance(columns, list):  # ✅ 특정 열에만 필터 적용
            col_indices = []
            header = [cell.value for cell in worksheet[1]]
            for col in columns:
                if col in header:
                    col_idx = header.index(col) + 1
                    col_indices.append(col_idx)

            if not col_indices:
                raise ValueError("지정한 컬럼이 시트에 존재하지 않습니다.")

            col_letters = [get_column_letter(idx) for idx in col_indices]
            ref_range = f"{col_letters[0]}1:{col_letters[-1]}{max_row}"
            worksheet.auto_filter.ref = ref_range

    @staticmethod
    @apply_to_cells
    def set_font(cell, font_name="Arial", font_size=12, bold=False, italic=False):
        font = Font(name=font_name, size=font_size, bold=bold, italic=italic)
        cell.font = font
        # for row in worksheet.iter_rows(min_row=1, max_row=1):
        #     for cell in row:
        #         cell.font = font

    @staticmethod
    @apply_to_cells
    def apply_border(cell, border_style="thin"):
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style),
        )
        cell.border = border
        # for row in worksheet.iter_rows():
        #     for cell in row:
        #         cell.border = border

    @staticmethod
    @apply_to_cells
    def apply_color(cell, color):
        """
        셀 배경색 적용
        - 16진수 색상 코드("FFFF00") 또는 색상 이름("red") 지원
        """
        # ✅ 색상 이름을 16진수로 변환
        if color.lower() in COLOR_MAP:
            color = COLOR_MAP[color.lower()]

        # ✅ 16진수 형식 보정
        if not color.startswith("#") and len(color) == 6:
            color = f"FF{color}"  # openpyxl은 ARGB 포맷을 사용

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.fill = fill
        # for row in worksheet.iter_rows():
        #     for cell in row:
        #         if cell.value:
        #             cell.fill = fill

    @staticmethod
    @apply_to_cells
    def set_alignment(cell, alignment="left"):
        """텍스트 정렬 (left, center, right)"""
        align_map = {"left": "left", "center": "center", "right": "right"}
        align = Alignment(horizontal=align_map.get(alignment, "left"))
        cell.alignment = align

    @staticmethod
    @apply_to_cells
    def set_number_format(cell, format_str):
        """숫자 형식 지정"""
        if isinstance(cell.value, (int, float)):
            cell.number_format = format_str

    # @staticmethod
    # def set_row_heights(worksheet, row_heights):
    #     """행 높이 설정"""
    #     if row_heights:
    #         for row, height in row_heights.items():
    #             worksheet.row_dimensions[row].height = height

    # @staticmethod
    # def highlight_rows(worksheet, row_colors):
    #     """특정 행 배경색 적용"""
    #     if row_colors:
    #         for row, color in row_colors.items():
    #             fill = PatternFill(
    #                 start_color=color, end_color=color, fill_type="solid"
    #             )
    #             for cell in worksheet[row]:
    #                 cell.fill = fill
