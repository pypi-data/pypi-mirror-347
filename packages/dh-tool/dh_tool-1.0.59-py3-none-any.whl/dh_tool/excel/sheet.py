import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


from .style import Style
from .dataframe_manager import DataFrameManager


class Sheet:
    def __init__(self, worksheet):
        self.worksheet = worksheet

    def write(self, data):
        if isinstance(data, dict):
            df = pd.DataFrame(data)
        elif isinstance(data, pd.DataFrame):
            df = data
        else:
            raise ValueError(
                "지원하지 않는 데이터 형식입니다. (dict 또는 DataFrame 사용 가능)"
            )
        df = self._convert_complex_types(df)
        for row in dataframe_to_rows(df, index=False, header=True):
            self.worksheet.append(row)
        return self

    def filter(self, include=None, exclude=None):
        df = self._get_dataframe_from_sheet()
        filtered_df = DataFrameManager.filter(df, include=include, exclude=exclude)
        self._refresh_worksheet(filtered_df)
        return self

    def aggregate(self, group_by, **aggregations):
        df = self._get_dataframe_from_sheet()
        aggregated_df = DataFrameManager.aggregate(df, group_by, **aggregations)
        self._refresh_worksheet(aggregated_df)
        return self

    def style(self, **kwargs):
        """
        다양한 엑셀 스타일을 적용합니다.

        ## ✅ 사용 가능한 옵션:
        - `auto_wrap` (bool): 텍스트 자동 줄바꿈 활성화
        - `freeze_first_row` (bool): 첫 번째 행 고정
        - `column_widths` (dict): 열 너비 설정 (예: {"A": 20, "name": 15})
        - `font` (dict): 폰트 설정
            - `name` (str): 폰트 이름 (기본값: "Arial")
            - `size` (int): 폰트 크기 (기본값: 12)
            - `bold` (bool): 굵은 글씨 여부 (기본값: False)
            - `italic` (bool): 이탤릭 여부 (기본값: False)
        - `border` (str): 테두리 스타일 ("thin", "medium", "dashed" 등)
        - `color` (str): 셀 배경색 (16진수 색상 코드, 예: "FFFF00")
        - `auto_adjust_columns` (bool): 데이터 길이에 따라 열 너비 자동 조정

        ## ✅ 사용 예제:
        sheet.style(
            auto_wrap=True,
            freeze_first_row=True,
            column_widths={"name": 20, "age": 15},
            font={"name": "Calibri", "size": 14, "bold": True},
            border="medium",
            color="FFFF00",
            auto_adjust_columns=True,
            filter=True,
        )
        """
        if kwargs.get("auto_wrap") and kwargs.get("auto_adjust_columns"):
            raise ValueError(
                "auto_wrap과 auto_adjust_columns는 함께 사용할 수 없습니다."
            )
        if kwargs.get("auto_wrap"):
            Style.apply_auto_wrap(self.worksheet)

        if kwargs.get("freeze_panes"):
            Style.freeze_panes(self.worksheet, kwargs["freeze_panes"])

        # if "column_widths" in kwargs:
        if kwargs.get("column_widths"):
            Style.set_column_width(self.worksheet, kwargs["column_widths"])

        if kwargs.get("auto_adjust_columns"):
            Style.auto_adjust_column_widths(self.worksheet)

        if kwargs.get("font"):
            font_options = kwargs["font"]
            Style.set_font(
                worksheet=self.worksheet,
                font_name=font_options.get("name", "Arial"),
                font_size=font_options.get("size", 12),
                bold=font_options.get("bold", False),
                italic=font_options.get("italic", False),
            )

        if kwargs.get("border"):
            Style.apply_border(self.worksheet, border_style=kwargs["border"])

        if kwargs.get("color"):
            Style.apply_color(self.worksheet, kwargs["color"])

        if kwargs.get("filter"):
            filter_columns = (
                kwargs["filter"] if isinstance(kwargs["filter"], list) else None
            )
            Style.apply_auto_filter(self.worksheet, filter_columns)

        if kwargs.get("number_format"):
            format_options = kwargs["number_format"]
            Style.set_number_format(self.worksheet, format_options)
        return self

    def style_to_cells(self, cells, **kwargs):
        """
        다양한 엑셀 스타일을 적용합니다.

        ## ✅ 사용 가능한 옵션:
        - `auto_wrap` (bool): 텍스트 자동 줄바꿈 활성화
        - `freeze_first_row` (bool): 첫 번째 행 고정
        - `column_widths` (dict): 열 너비 설정 (예: {"A": 20, "name": 15})
        - `font` (dict): 폰트 설정
            - `name` (str): 폰트 이름 (기본값: "Arial")
            - `size` (int): 폰트 크기 (기본값: 12)
            - `bold` (bool): 굵은 글씨 여부 (기본값: False)
            - `italic` (bool): 이탤릭 여부 (기본값: False)
        - `border` (str): 테두리 스타일 ("thin", "medium", "dashed" 등)
        - `color` (str): 셀 배경색 (16진수 색상 코드, 예: "FFFF00")
        - `auto_adjust_columns` (bool): 데이터 길이에 따라 열 너비 자동 조정

        ## ✅ 사용 예제:
        sheet.style(
            auto_wrap=True,
            freeze_first_row=True,
            column_widths={"name": 20, "age": 15},
            font={"name": "Calibri", "size": 14, "bold": True},
            border="medium",
            color="FFFF00",
            auto_adjust_columns=True,
            filter=True,
        )
        """
        if kwargs.get("auto_wrap") and kwargs.get("auto_adjust_columns"):
            raise ValueError(
                "auto_wrap과 auto_adjust_columns는 함께 사용할 수 없습니다."
            )
        if kwargs.get("auto_wrap"):
            Style.apply_auto_wrap(self.worksheet)

        if kwargs.get("freeze_panes"):
            Style.freeze_panes(self.worksheet, kwargs["freeze_panes"])

        # if "column_widths" in kwargs:
        if kwargs.get("column_widths"):
            Style.set_column_width(self.worksheet, kwargs["column_widths"])

        if kwargs.get("auto_adjust_columns"):
            Style.auto_adjust_column_widths(self.worksheet)

        if kwargs.get("font"):
            font_options = kwargs["font"]
            Style.set_font(
                worksheet=self.worksheet,
                font_name=font_options.get("name", "Arial"),
                font_size=font_options.get("size", 12),
                bold=font_options.get("bold", False),
                italic=font_options.get("italic", False),
                cells=cells,
            )

        if kwargs.get("border"):
            Style.apply_border(
                self.worksheet, border_style=kwargs["border"], cells=cells
            )

        if kwargs.get("color"):
            Style.apply_color(self.worksheet, kwargs["color"], cells=cells)

        if kwargs.get("filter"):
            filter_columns = (
                kwargs["filter"] if isinstance(kwargs["filter"], list) else None
            )
            Style.apply_auto_filter(self.worksheet, filter_columns)

        return self

    # ✅ 1. 셀에 하이퍼링크 추가
    def add_hyperlink(self, cell, url, display=None):
        """특정 셀에 하이퍼링크 추가"""
        current_value = self.worksheet[cell].value  # 기존 셀 값 저장
        self.worksheet[cell].hyperlink = url
        self.worksheet[cell].value = display if display is not None else current_value
        self.worksheet[cell].style = "Hyperlink"
        return self

    # ✅ 2. 컬럼 전체에 하이퍼링크 추가
    def add_hyperlinks_to_column(self, column_name, urls, display_texts=None):
        """컬럼의 각 셀에 하이퍼링크 추가"""
        df = self._get_dataframe_from_sheet()

        if len(urls) != len(df):
            print(len(urls), len(df))
            raise ValueError("URL 리스트의 길이가 데이터프레임과 일치해야 합니다.")

        if display_texts and len(display_texts) != len(df):
            raise ValueError(
                "표시 텍스트 리스트의 길이가 데이터프레임과 일치해야 합니다."
            )

        col_idx = df.columns.get_loc(column_name) + 1
        for i, url in enumerate(
            urls, start=2
        ):  # start=2 → 헤더를 제외하고 2행부터 시작
            cell = f"{get_column_letter(col_idx)}{i}"  # A, B, C 열 등으로 변환
            # cell = f"{chr(64 + col_idx)}{i}"  # A, B, C 열 등으로 변환
            display = display_texts[i - 2] if display_texts else None
            self.add_hyperlink(cell, url, display)

        return self

    def _refresh_worksheet(self, df):
        self.worksheet.delete_rows(1, self.worksheet.max_row)
        for row in dataframe_to_rows(df, index=False, header=True):
            self.worksheet.append(row)

    def _get_dataframe_from_sheet(self):
        data = list(self.worksheet.values)
        headers = data[0]
        rows = data[1:]
        return pd.DataFrame(rows, columns=headers)

    def _convert_complex_types(self, df):
        """리스트나 딕셔너리 데이터를 문자열로 변환"""
        for col in df.columns:
            if df[col].apply(lambda x: isinstance(x, (list, dict))).any():
                df[col] = df[col].apply(
                    lambda x: str(x) if isinstance(x, (list, dict)) else x
                )
        return df
