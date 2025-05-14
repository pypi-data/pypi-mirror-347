# -*- coding: utf-8 -*-
"""
This module offers a set of operations that user can modify their excel files.
"""
import os
import pathlib
import warnings
from importlib.resources import files
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from urllib3 import disable_warnings

from .excel_definition import *
from .jira_client import *
from .milestone import *
from .priority import *
from .sprint_schedule import *
from .story import *

__all__ = [
    "read_excel_file",
    "output_to_excel_file",
    "output_to_csv_file",
    "process_excel_file",
]

# Currently, the openpyxl package will report an obsolete warning.
warnings.simplefilter(action="ignore", category=UserWarning)
# Disable the HTTPS certificate verification warning.
disable_warnings()

HERE = pathlib.Path(__file__).resolve().parent
ASSETS = HERE / "assets"


def read_excel_file(
    file: "str | Path",
    excel_definition: ExcelDefinition,
    sprint_schedule: SprintScheduleStore,
) -> tuple[list[str], list[Story]]:
    """
    Read and parse the excel file

    :parm file:
        The excel file that you want to read

    :parm excel_definition:
        The excel column definition which is imported from the :py:class:`ExcelDefinition`

    :parm sprint_schedule:
        The priority mapping for the :py:class:`Milestone` object.

    :return:
        A :py:class:`tuple` object which contains a list of column name and a list of :py:class:`Story`.
    """

    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError(f"The input excel file is invalid.")

    if not pathlib.Path(file).exists():
        raise ValueError(f"The input excel file: {file} cannot be found.")

    wb: Workbook = openpyxl.load_workbook(
        str(file), read_only=True, keep_vba=False, data_only=True, keep_links=True
    )

    if wb.active is None or (
        type(wb.active) is not Worksheet and type(wb.active) is not ReadOnlyWorksheet
    ):
        raise ValueError("The input excel file doesn't contain any sheets.")

    sheet: Worksheet = wb.active

    # The count of column is taking from the definition file to avoid too many columns inside the excel file.
    # Also, need to avoid exceed the range of the actual count.
    if excel_definition.max_column_index > sheet.max_column:
        column_count = sheet.max_column
    else:
        column_count = excel_definition.max_column_index

    max_column_index = chr(65 + column_count - 1)

    start_column = "A1"
    end_column = f"{max_column_index}1"
    columns: list[str] = []

    for cells in sheet[start_column:end_column]:
        if len(columns) > 0:
            break
        for cell in cells:
            # TODO: Copy styles
            columns.append(cell.value)

    start_cell = "A2"
    end_cell = f"{max_column_index}{sheet.max_row}"
    rows = sheet[start_cell:end_cell]

    stories = []

    excel_definition_columns = excel_definition.get_columns()
    storyFactory = StoryFactory(excel_definition_columns)

    for row in rows:
        if _should_skip(row):
            continue

        story: Story = storyFactory.create_story()
        for column_index in range(len(row)):
            column = excel_definition_columns[column_index]
            if column["name"] is None:
                continue
            story.set_value(column["type"], column["name"], row[column_index].value)

        story.calc_sprint_schedule(sprint_schedule)
        stories.append(story)

    wb.close()
    return (columns, stories)


def _should_skip(row: list) -> bool:
    if len(row) == 0:
        return True
    else:
        first_cell_value = row[0].value
        if first_cell_value is None or len(str(first_cell_value)) == 0:
            return True
    return False


def output_to_csv_file(
    file: str,
    stories: "list[Story]",
    over_write: bool = True,
):
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError(f"The file is invalid.")

    if not pathlib.Path(file).exists():
        if over_write is True:
            os.remove(file)
        else:
            raise ValueError(f"The csv file: {file} is already exist.")

    with open(file, mode="w") as csv_file:
        separator = "-" * 300
        for story in stories:
            csv_file.write(f"{separator}\n")
            csv_file.write(str(story))


def output_to_excel_file(
    file: "str | Path",
    stories: "list[Story]",
    excel_definition: ExcelDefinition,
    columns_in_excel: "list[str] | None" = None,
    over_write: bool = True,
):
    """
    Generate excel file

    :parm file:
        Output excel file name including the path

    :parm stories:
        A list of :py:class:`Story` which need to be wrote to the excel

    :parm excel_definition:
        The excel column definition which is imported from the :py:class:`ExcelDefinition`

    :parm columns_in_excel:
        Using separate column names instead of importing from the :py:class:`ExcelDefinition`. Usually, it comes from the input excel file.

    :parm over_write:
        Whether or not the exist output file will be over-write.
    """
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError(f"The output file name is invalid.")

    if pathlib.Path(file).exists():
        if over_write is True:
            try:
                os.remove(file)
            except PermissionError as e:
                print(e)
                return
        else:
            raise ValueError(f"The output excel file: {file} is already exist.")

    wb = openpyxl.Workbook()

    if wb.active is None or type(wb.active) is not Worksheet:
        raise ValueError("The output file doesn't contain any active sheets.")

    sheet: Worksheet = wb.active

    # TODO: Refactor this method to return needed properties list.
    excel_definition_columns = excel_definition.get_columns()

    # Use original excel column name first.
    columns = columns_in_excel
    if columns is None:
        columns = [column["name"] for column in excel_definition_columns]

    for column_index in range(len(columns)):
        cell = sheet.cell(row=1, column=column_index + 1)
        # There are three kinds of Cells. Only the Cell has the value attribute.
        if hasattr(cell, "value"):
            setattr(cell, "value", columns[column_index])

    if stories is not None:
        for row_index in range(len(stories)):
            for column in excel_definition_columns:
                if column["name"] is None:
                    continue
                cell = sheet.cell(row=row_index + 2, column=column["index"])
                if hasattr(cell, "value"):
                    setattr(
                        cell, "value", stories[row_index].format_value(column["name"])
                    )

    wb.save(str(file))
    wb.close()


def _query_jira_information(stories: list[Story], excel_definition: ExcelDefinition):
    from dotenv import load_dotenv

    load_dotenv(ASSETS / ".env")

    jira_url: str | None = os.environ.get("JIRA_URL", default=None)
    if (
        # TODO: Avoid duplicate logic.
        jira_url is None
        or jira_url.isspace()
        or len(jira_url) == 0
    ):
        print(
            "The jira url is invalid. Please use the update-jira-info command to add/update url."
        )
        return

    jira_acccess_token: str | None = os.environ.get("JIRA_ACCESS_TOKEN", default=None)
    if (
        jira_acccess_token is None
        or jira_acccess_token.isspace()
        or len(jira_acccess_token) == 0
    ):
        print(
            "The jira access token is invalid. Please use the update-jira-info command to add/update token."
        )
        return

    jira_client = JiraClient(jira_url, jira_acccess_token)

    if not jira_client.health_check():
        print(
            "The jira access token is revoked. Please use the update-jira-info command to add/update token."
        )
        return

    jira_fields = []

    for definition_column in excel_definition.get_columns():
        if (
            definition_column["name"] is None
            or definition_column["jira_field_mapping"] is None
        ):
            continue
        else:
            jira_fields.append(
                {
                    "name": definition_column["name"],
                    "jira_name": definition_column["jira_field_mapping"]["name"],
                    "jira_path": definition_column["jira_field_mapping"]["path"],
                }
            )

    jira_query_result = jira_client.get_stories_detail(
        [story["storyId"].strip() for story in stories], jira_fields
    )

    for story in stories:
        story_id: str = story["storyId"].lower().strip()
        if story_id in jira_query_result:
            for jira_field in jira_fields:
                story[jira_field["name"]] = jira_query_result[story_id][
                    jira_field["jira_name"]
                ]
        else:
            # Story ID has been changed because of convertion.
            temp_result = jira_client.get_stories_detail([story_id], jira_fields)
            if len(temp_result) > 0:
                story["storyId"] = list(temp_result.keys())[0].upper()
                for jira_field in jira_fields:
                    story[jira_field["name"]] = list(temp_result.values())[0][
                        jira_field["jira_name"]
                    ].upper()
                print(
                    f"Story id has been changed. Previous: {story_id.upper()}, Current: {story['storyId'].upper()}"
                )
            else:
                print(f"Cannot find related information for story: {story_id}")
                story.need_sort = False  # TODO: Avoid NoneType compare issue.
                continue


def process_excel_file(
    input_file: "str | Path",
    output_file: "str | Path",
    excel_definition_file: "str | Path | None" = None,
    sprint_schedule_file: "str | Path | None" = None,
    over_write: bool = True,
):
    """
    Sort the excel file and output the result

    :parm input_file:
        The excel file need to be sorted. (Absolute path only)

    :parm output_file:
        The sorted excel file location. (Absolute path only)

    :parm sprint_schedule_file:
        The JSON file which contains the priority list to calculate the :py:class:`Milestone`

    :parm excel_definition_file:
        The JSON file which contains the input excel file's structure.

    :parm over_write:
        Whether or not the exist output file will be over-write.
    """
    sprint_schedule = SprintScheduleStore()
    if sprint_schedule_file is None:
        sprint_schedule.load(
            files("jira_tool.assets").joinpath("sprint_schedule.json").read_text()
        )
        print("Using default sprint schedule...")
    else:
        sprint_schedule.load_file(sprint_schedule_file)
        print("Using custom sprint schedule...")

    excel_definition = ExcelDefinition()
    if excel_definition_file is None:
        excel_definition.load(
            files("jira_tool.assets").joinpath("excel_definition.json").read_text()
        )
        print("Using default excel definition...")
    else:
        excel_definition.load_file(excel_definition_file)
        print("Using custom excel definition...")

    validation_result = excel_definition.validate()
    if len(validation_result) != 0:
        print(
            "Validating excel definition failed. Please check below information to fix first."
        )
        for item in validation_result:
            print(item)
        return
    else:
        print("Validating excel definition success.")

    excel_columns, stories = read_excel_file(
        input_file, excel_definition, sprint_schedule
    )

    if stories is None:
        print("There are no stories inside the excel file.")
        return

    # Execute pre-process steps
    pre_process_steps = excel_definition.get_pre_process_steps()

    for pre_process_step in pre_process_steps:
        print(f"Executing step: {pre_process_step['name']}...")
        if pre_process_step["name"].lower() in "RetrieveJiraInformation".lower():
            need_call_jira_api: bool = False
            for excel_definition_column in excel_definition.get_columns():
                if excel_definition_column["jira_field_mapping"] is not None:
                    need_call_jira_api = True
                    break

            if need_call_jira_api:
                stories_need_call_jira: list[Story] = []
                for story in stories:
                    if story.need_sort:
                        stories_need_call_jira.append(story)
                _query_jira_information(stories_need_call_jira, excel_definition)
        elif pre_process_step["name"].lower() in "FilterOutStoryWithoutId".lower():
            for story in stories:
                if story["storyId"] is None:
                    story.need_sort = False
        elif (
            pre_process_step["name"].lower()
            in "FilterOutStoryBasedOnJiraStatus".lower()
        ):
            for story in stories:
                if story["status"] is not None and story[
                    "status"
                ].upper() in pre_process_step["config"].get("JiraStatuses", []):
                    story.need_sort = False
        print(f"Executing finish.")

    stories_no_need_sort = []
    stories_need_sort = []

    for story in stories:
        if story.need_sort:
            stories_need_sort.append(story)
        else:
            stories_no_need_sort.append(story)

    # Execute sorting logic.
    sort_strategies = excel_definition.get_sort_strategies()

    for sort_strategy in sort_strategies:
        print(f"Executing {sort_strategy['name']} sorting...")
        if sort_strategy["name"] is None:
            continue
        elif sort_strategy["name"].lower() in "InlineWeights".lower():
            stories_need_sort = sorted(stories_need_sort, reverse=True)
        elif sort_strategy["name"].lower() in "SortOrder".lower():
            sort_stories_by_property_and_order(
                stories_need_sort, excel_definition, sort_strategy["config"]
            )
        elif sort_strategy["name"].lower() in "RaiseRanking".lower():
            stories_need_sort = sort_stories_by_raise_ranking(
                stories_need_sort, excel_definition, sort_strategy["config"]
            )
        print(f"Executing finish.")

    output_to_excel_file(
        output_file,
        stories_need_sort + stories_no_need_sort,  # First output the sorted stories.
        excel_definition,
        excel_columns,
        over_write,
    )

    print(f"{output_file} has been saved.")
